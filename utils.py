"""Main functions to be utilized throughout
the project can be found here."""
import copy
import csv
import os
import json
import random
import string
import io
import tempfile
from typing import Union
import pandas as pd
import xlsxwriter
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook


def calculator(
    excel_document,
    excel_sheet,
    projekt_tid,
    kalkylrantan,
    skattesats,
    grundinvestering,
    utbet_ar_noll,
    rorelsebindandekapital,
    avskrivningar,
    inbet,
    utbet,
    rest,
    acc_list
):
    """Main functionality of the program calculates all components
    based on given parameters and creates an excel portfolio"""

    alfabet = string.ascii_uppercase
    count = 1

    # Create formats
    bold_format = excel_document.add_format({"bold": True})
    italic_format = excel_document.add_format({"italic": True})
    bold_center__color_format = excel_document.add_format({
        "bold": True,
        "align": "center"
    })
    center_economic_format = excel_document.add_format({
        "align": "center",
        "num_format": "#,##0.000"
    })
    bold_center_econ_format = excel_document.add_format({
        "bold": True,
        "align": "center",
        "num_format": "#,##0.00"
    })
    percent_format = excel_document.add_format({
        "italic": True,
        "bold": True,
        "num_format": "0.0%"
    })

    # Set column widths
    excel_sheet.set_column("A:A", 25)
    excel_sheet.set_column("B:X", 15, center_economic_format)

    # Headlines
    excel_sheet.write("A1", "Project lifetime", bold_format)
    excel_sheet.write("A2", "Initial Investment")
    excel_sheet.write("A3", "Depreciation")
    excel_sheet.write("A4", "Incoming Payments")
    excel_sheet.write("A5", "Outgoing Payments")
    excel_sheet.write("A6", "Residual")
    excel_sheet.write("A7", "Restricted Equity")
    excel_sheet.write("A8", "Yearly Net")
    excel_sheet.write("A9", "Present Value")
    excel_sheet.write("A10", "Accumulated Present Value")
    excel_sheet.write("A11", "Net Present Value", bold_format)
    excel_sheet.write("A13", "Nominal Discount Rate", italic_format)
    excel_sheet.write("A14", "Discount Rate After TaX", italic_format)
    excel_sheet.write("A15", "Tax Rate", italic_format)

    # Column headers for years
    for letter in alfabet[1:projekt_tid + 2]:
        excel_sheet.write(f"{letter}1", count - 1, bold_center__color_format)
        count += 1

    # Initial investment
    excel_sheet.write("B2", grundinvestering)

    # Depreciation
    for letter in alfabet[2:projekt_tid + 2]:
        excel_sheet.write(f"{letter}3", avskrivningar)

    # Incoming payments
    for letter in alfabet[2:projekt_tid + 2]:
        excel_sheet.write(f"{letter}4", inbet)

    # Outgoing payments
    excel_sheet.write("B5", utbet_ar_noll)
    for letter in alfabet[2:projekt_tid + 2]:
        excel_sheet.write(f"{letter}5", utbet)

    # Residual and restricted equity
    excel_sheet.write(f"{alfabet[projekt_tid + 1]}6", rest)
    excel_sheet.write("B7", rorelsebindandekapital)
    excel_sheet.write(f"{alfabet[projekt_tid + 1]}7", -rorelsebindandekapital)

    # Net values
    ar_noll_netto = grundinvestering + utbet_ar_noll + rorelsebindandekapital
    ack_mellan_nuvarde = ar_noll_netto
    excel_sheet.write("B8", ar_noll_netto, bold_center_econ_format)

    ar_mellan_netto = inbet + utbet + avskrivningar
    for letter in alfabet[2:projekt_tid + 1]:
        excel_sheet.write(
            f"{letter}8", ar_mellan_netto, bold_center_econ_format
        )

    ar_sista_netto = (
        inbet + utbet + avskrivningar - rorelsebindandekapital + rest
    )
    excel_sheet.write(
        f"{alfabet[projekt_tid + 1]}8", ar_sista_netto, bold_center_econ_format
    )

    # Present value
    excel_sheet.write("B9", ar_noll_netto)
    counter = 1
    for letter in alfabet[2:projekt_tid + 1]:
        val = ar_mellan_netto / ((1 + kalkylrantan) ** counter)
        excel_sheet.write(f"{letter}9", val)
        counter += 1

    ar_sista_nuvarde = ar_sista_netto / ((1 + kalkylrantan) ** projekt_tid)
    excel_sheet.write(f"{alfabet[projekt_tid + 1]}9", ar_sista_nuvarde)

    # Accumulated present value
    excel_sheet.write("B10", ar_noll_netto)
    acc_list.append(ar_noll_netto)
    counter_ett = 1
    for letter in alfabet[2:projekt_tid + 1]:
        val = ar_mellan_netto / ((1 + kalkylrantan) ** counter_ett)
        ack_mellan_nuvarde += val
        acc_list.append(ack_mellan_nuvarde)
        excel_sheet.write(f"{letter}10", ack_mellan_nuvarde)
        counter_ett += 1

    ar_sista_ack_nuvarde = ack_mellan_nuvarde + ar_sista_nuvarde
    acc_list.append(ar_sista_ack_nuvarde)
    excel_sheet.write(f"{alfabet[projekt_tid + 1]}10", ar_sista_ack_nuvarde)

    # Final summary values
    excel_sheet.write("B11", ar_sista_ack_nuvarde, bold_center_econ_format)
    excel_sheet.write("B13", kalkylrantan / (1 - skattesats), percent_format)
    excel_sheet.write("B14", kalkylrantan, percent_format)
    excel_sheet.write("B15", skattesats, percent_format)

    return ar_sista_ack_nuvarde, acc_list, avskrivningar


def colorizer(filename, project):
    """
    Applies color to distinguish positive and negative values
    in a given worksheet of an Excel file.

    Arguments:
    - filename (str): Full name of the Excel file, e.g. "Alice Portfolio.xlsx"
    - project (str): Name of the worksheet, e.g. "Projekt1 Portfolio.xlsx"
    """

    # Load the workbook and select the specific sheet
    wb = load_workbook(filename)
    if project not in wb.sheetnames:
        raise ValueError(f"Worksheet '{project}' not found in '{filename}'")

    ws = wb[project]

    # Define the colors for positive and negative values
    green_fill = PatternFill(
        start_color='FF00FF00', end_color='FF00FF00', fill_type='solid'
    )
    red_fill = PatternFill(
        start_color='FFFF0000', end_color='FFFF0000', fill_type='solid'
    )

    # Iterate through each cell (excluding headers)
    for row in ws.iter_rows(
        min_row=2, max_row=12, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value > 0:
                    cell.fill = green_fill
                elif cell.value < 0:
                    cell.fill = red_fill

    wb.save(filename)


def excel_merger(
    excel_portfolio_name,
    excel_proj,
    write_excel_content_fn,
    temp_filename="__temp_project_sheet.xlsx"
):
    """
    Creates a sheet in a temporary Excel file using xlsxwriter,
    then merges that sheet into the real portfolio file using openpyxl.

    Parameters:
    - EXCEL_PORTFOLIO_NAME (str): target Excel file to merge the sheet into
    - EXCEL_PROJ (str): name of the sheet (project)
    - write_excel_content_fn (function): takes (workbook, worksheet) input data
    - temp_filename (str): name of the temp file to use
    """
    # Step 1: Create the sheet in a temp file with xlsxwriter
    excel_document = xlsxwriter.Workbook(temp_filename)
    excel_sheet = excel_document.add_worksheet(excel_proj)

    # Call the user-defined function to write content
    ar_sista_ack_nuvarde, acc_list, avskrivningar = (
        write_excel_content_fn(excel_document, excel_sheet)
    )
    excel_document.close()

    # Step 2: Copy sheet contents to the real workbook using openpyxl
    src_wb = load_workbook(temp_filename)
    src_ws = src_wb.active  # Assume it's the only sheet

    target_wb = load_workbook(excel_portfolio_name)

    if excel_proj in target_wb.sheetnames:
        print(
            f"Sheet '{excel_proj}' already exists in '{excel_portfolio_name}'."
        )
    else:
        dest_ws = target_wb.create_sheet(excel_proj)
        for row in src_ws.iter_rows():
            for cell in row:
                dest_ws[cell.coordinate].value = cell.value
        target_wb.save(excel_portfolio_name)
        print(f"Added sheet '{excel_proj}' to '{excel_portfolio_name}'.")

    # Clean up temp file
    src_wb.close()
    os.remove(temp_filename)

    return ar_sista_ack_nuvarde, acc_list, avskrivningar


def extract_values(project_df, year_columns, title):
    """Extracts float values for a given object title across year columns."""
    row = project_df[project_df["object title"] == title]
    if row.empty:
        return [0.0] * len(year_columns)
    return row.iloc[0][year_columns].fillna(0).astype(float).tolist()


def file_path_creator() -> str:
    """handles the conditions for githubs testing environment,
    filepath creator."""
    if os.getenv('GITHUB_ACTIONS') == 'true':
        # GitHub Actions environment
        local_path = (
            "https://github.com/ProDevOperationsEngineer/"
            "Investmentcalculator/blob/main/calculator.py"
        )
    else:
        # Local environment
        local_path = "calculator.py"
    return local_path


def generate_random_id(length=8):
    """Generate a random string of fixed length."""
    letters_and_digits = string.ascii_letters + string.digits
    return ''.join(random.choice(letters_and_digits) for i in range(length))


def generate_full_portfolio(csv_filename, username):
    """
    Creates a complete investment portfolio workbook for a user by combining
    all of their projects (each as a worksheet) into one Excel file.
    """
    df = pd.read_csv(csv_filename, dtype=object)
    user_projects = df[df["username"] == username]["project"].unique()

    if len(user_projects) == 0:
        return None

    # Create final master workbook
    master_wb = Workbook()
    master_wb.remove(master_wb.active)  # remove default empty sheet

    for project in user_projects:
        # Generate Excel file in memory for this project
        project_stream = load_from_csv_excel(csv_filename, username, project)
        if project_stream is None:
            continue

        # Save to temp file for colorizer
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_filename = tmp.name
            with open(tmp_filename, "wb") as f:
                f.write(project_stream.read())

        # Apply coloring
        colorizer(tmp_filename, project)

        # Load and copy the sheet into master workbook
        temp_wb = load_workbook(tmp_filename)
        temp_ws = temp_wb[project]

        # Create new sheet in master workbook
        new_ws = master_wb.create_sheet(title=project)

        for col_idx in range(1, temp_ws.max_column + 1):
            letter = get_column_letter(col_idx)
            src_dim = temp_ws.column_dimensions.get(letter)
            if src_dim and src_dim.width:
                new_ws.column_dimensions[letter].width = src_dim.width
            else:
                # fallback width if not defined
                new_ws.column_dimensions[letter].width = 15

        for row in temp_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(
                    row=cell.row, column=cell.col_idx, value=cell.value
                )
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = copy.copy(cell.alignment)

        temp_wb.close()
        os.remove(tmp_filename)

    # Save final master workbook to stream
    output = io.BytesIO()
    master_wb.save(output)
    output.seek(0)
    return output


def get_last_user(data: list, username: str) -> dict | None:
    """Returns the last user entry in the list that matches the username."""
    for entry in reversed(data):  # go from the end backwards
        if entry.get("username") == username:
            return entry
    return None  # if no match found


def json_file_amender(filename, investor_data) -> list:
    """Amends JSON file by overwriting projects with the same name."""

    # Load or init
    if os.path.exists(filename):
        with open(filename, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError:
                data = []
    else:
        data = []

    username = investor_data.get("username")

    if data:
        latest_entry = data[-1]

        if latest_entry.get("username") == investor_data.get("username"):
            existing_projects = latest_entry["projects"]
            new_projects = investor_data["projects"]

            for new_project in new_projects:
                # Remove any existing project with the same name
                existing_projects[:] = [
                    p for p in existing_projects
                    if p.get("project_name") != new_project.get("project_name")
                ]
                # Add the new version
                existing_projects.append(new_project)
        else:
            # Remove all previous entries with the same username
            data = (
                [entry for entry in data if entry.get("username") != username]
            )
            data.append(investor_data)
    else:
        data.append(investor_data)

    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)

    return data


def load_from_csv_excel(csv_filename, username, project):
    """
    Loads a specific project from a row-based CSV database,
    generates a formatted Excel sheet using calculator(),
    applies cell coloring via colorizer(),
    and returns the result as a downloadable stream.
    """
    # Read the full database
    df = pd.read_csv(csv_filename, dtype=object)

    # Filter for user + project
    project_df = df[(df["username"] == username) & (df["project"] == project)]
    if project_df.empty:
        return None

    # Detect which columns are years (0, 1, ..., N)
    year_columns = [col for col in df.columns if col.isdigit()]
    projekt_tid = len(year_columns) - 1  # exclude year 0

    # Extract a specific variable’s values as a list of floats
    def get_values(title):
        row = project_df[project_df["object title"] == title]
        if row.empty:
            return [0.0] * len(year_columns)
        return row.iloc[0][year_columns].fillna(0).astype(float).tolist()

    # Extract fields required by your calculator()
    grundinvestering = get_values("Initial Investment")[0]
    avskrivningar_list = get_values("Depreciation")
    inbet_list = get_values("Incoming Payments")
    utbet_list = get_values("Outgoing Payments")
    rest = get_values("Residual")[-1]
    rorelsebindandekapital = get_values("Restricted Equity")[0]
    kalkylrantan = get_values("Discount Rate After TaX")[0]
    skattesats = get_values("Tax Rate")[0]
    acc_list = []

    # Create Excel file in memory
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet(project)

    # Generate content using your calculator
    calculator(
        workbook,
        worksheet,
        projekt_tid,
        kalkylrantan,
        skattesats,
        grundinvestering,
        utbet_list[0],              # utbet_ar_noll
        rorelsebindandekapital,
        avskrivningar_list[1],
        inbet_list[1],
        utbet_list[1],
        rest,
        acc_list
    )

    # Finish writing in memory
    workbook.close()

    # Save to a temp file so colorizer (which expects a file) can process it
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_filename = tmp.name
        with open(tmp_filename, "wb") as f:
            f.write(output.getvalue())

    # Run the colorizer on the temp file
    colorizer(tmp_filename, project)

    # Read the final version back into memory
    with open(tmp_filename, "rb") as f:
        final_bytes = f.read()

    # Clean up the temp file
    os.remove(tmp_filename)

    # Return the colored Excel file as a stream
    return io.BytesIO(final_bytes)


def load_from_csv(filename) -> list:
    """Loads data from a CSV file and returns a list of dictionaries."""
    data_list = []
    with open(filename, mode="r", newline='', encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            data_list.append(row)
    return data_list


def load_from_csv_image(filename) -> list:
    """Loads image data from a CSV file and returns a list of dictionaries."""
    data_list = []
    with open(filename, mode="r", newline="", encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            data_list.append(row)
    return data_list


def load_last_shared_data() -> dict:
    """Loads the last dictionary from shared data,
    creating the file if it doesn't exist."""
    file_path = "shared_data.json"

    # Create the file if it doesn't exist
    if not os.path.exists(file_path):
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump({}, f)

    # Check if the file is non-empty
    if os.path.getsize(file_path) > 0:
        with open(file_path, 'r', encoding='utf-8') as f:
            shared_data_list = json.load(f)
            if shared_data_list:
                return shared_data_list[-1]
            else:
                print("Warning: File contains an empty list.")
    else:
        print("Warning: File is empty.")

    return {}  # Return an empty dict if no valid data is found


def load_shared_data() -> list:
    """Loads the dictionary from shared data,
    creating the file if it doesn't exist."""
    file_path = "shared_data.json"

    if not os.path.exists(file_path):
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump([], f)

    if os.path.getsize(file_path) == 0:
        return []

    with open(file_path, 'r', encoding='utf-8') as f:
        try:
            data = json.load(f)
            if isinstance(data, list):
                return data
            else:
                print("Warning: shared_data.json is not a list.")
                return []
        except json.JSONDecodeError:
            print("Warning: shared_data.json is corrupted.")
            return []


def save_to_csv_excel(
    excel_filename,
    project,
    csv_filename,
    username=None
):
    """
    Converts a project sheet from an Excel file to CSV, preserving structure
    but rounding all numeric values to whole integers.
    """
    if not os.path.exists(excel_filename):
        raise FileNotFoundError(
            f"Excel file '{excel_filename}' does not exist."
        )

    try:
        df = pd.read_excel(excel_filename, sheet_name=project, dtype=object)
    except ValueError as exc:
        raise ValueError(
            f"Sheet '{project}' not found in '{excel_filename}'."
        ) from exc

    # Add metadata columns (optional, but reversible)
    df.insert(0, "project", project)
    if username:
        df.insert(0, "username", username)

    if os.path.exists(csv_filename):
        columns = ["project"]
        if username:
            columns.append("username")

        existing_df = pd.read_csv(csv_filename, usecols=columns)

        if username:
            match = (
                (
                    existing_df["project"] == project
                ) & (
                    existing_df["username"] == username)
            ).any()
        else:
            match = (existing_df["project"] == project).any()
        if match:
            print(
                f" Project '{project}' already exists in '{csv_filename}'"
            )
            os.remove(excel_filename)
            print(f"Deleted Excel file '{excel_filename}'")
            return  # Skip writing

    # Fix header label if needed
    if len(df.columns) > 2 and df.columns[2] != "object title":
        df.columns.values[2] = "object title"

    # Round all numeric values to nearest integer (preserve labels and headers)
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            df[col] = df[col].round(0).astype("Int64")

    # Append to CSV
    file_exists = os.path.exists(csv_filename)
    df.to_csv(csv_filename, mode="a", header=not file_exists, index=False)

    print(f" Saved sheet '{project}' to '{csv_filename}'")

    # Delete the source Excel file to save memory (optional)
    os.remove(excel_filename)
    print(f" Deleted Excel file '{excel_filename}'")


def save_to_csv_image(data, csv_filename, mode="a"):
    """Saves image in byte64data to a CSV file"""
    file_exists = os.path.isfile(csv_filename)

    if file_exists:
        project_list = load_from_csv(csv_filename)
        project_exist = 0
        for projects in project_list:
            if projects["project_name"] == data["project_name"]:
                project_exist = 1
        if project_exist == 0:
            with open(
                csv_filename, mode, newline='', encoding="utf-8"
            ) as csv_file:
                writer = csv.writer(csv_file)
                if not file_exists:
                    writer.writerow(
                        ["username", "project_name", "Base64Data"]
                    )
                writer.writerow(data.values())
    else:
        with open(
            csv_filename, mode, newline='', encoding="utf-8"
        ) as csv_file:
            writer = csv.writer(csv_file)
            if not file_exists:
                writer.writerow(["username", "project_name", "Base64Data"])
            writer.writerow(data.values())


def save_to_csv_project(data, filename):
    """Saves or updates a project in a CSV file."""
    fieldnames = [
        "username", "project_name", "lifetime", "initial_investment",
        "incoming_payments", "outgoing_payments", "outgoing_payments_0",
        "restricted_equity", "residual", "discount_rate", "tax_rate",
        "net_present_value", "depreciation",
    ]

    data = {key: data.get(key, "") for key in fieldnames}

    project_list = []
    updated = False

    if os.path.isfile(filename):
        with open(filename, mode="r", newline='', encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                if row["project_name"] == data["project_name"]:
                    project_list.append(data)
                    updated = True
                else:
                    project_list.append(row)
    else:
        project_list = [data]
        updated = True

    if not updated:
        project_list.append(data)

    with open(filename, mode="w", newline='', encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(project_list)


def save_to_csv_user(data, filename, mode='a'):
    """Saves a dictionary to a CSV file, optionally in append mode"""

    fieldnames = ["username", "password"]

    file_exists = os.path.isfile(filename)

    if file_exists:
        project_list = load_from_csv(filename)
        user_exist = 0
        for projects in project_list:
            if projects["username"] == data["username"]:
                user_exist = 1
        if user_exist == 0:
            with open(
                filename, mode, newline='', encoding="utf-8"
            ) as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writerow(data)
    else:
        with open(
            filename, mode, newline='', encoding="utf-8"
        ) as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerow(data)


def taxes(
    inb: Union[int, float],
    utb: Union[int, float],
    utb_ar_noll: Union[int, float],
    re: Union[int, float],
    kalk: Union[int, float],
    skatt: Union[int, float]
) -> tuple[
    Union[int, float],
    Union[int, float],
    Union[int, float],
    Union[int, float],
    Union[int, float],
]:
    """Modifying the values to be accounting for taxes and
    returns the new values to be assigned to the old variables
    """
    inb *= (1-skatt)
    utb *= (1-skatt)
    utb_ar_noll *= (1-skatt)
    kalk *= (1-skatt)
    re *= (1-skatt)
    return inb, utb, utb_ar_noll, re, kalk
