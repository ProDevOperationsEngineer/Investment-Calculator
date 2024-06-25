"""Corrects negative values to appear in red color
and positive in green
"""


import openpyxl
from openpyxl.styles import PatternFill


# Load the workbook and select the active worksheet
wb = openpyxl.load_workbook("InvestmentCalc.xlsx")
ws = wb.active

# Define the colors for positive and negative values
green_fill = PatternFill(
    start_color='FF00FF00', end_color='FF00FF00', fill_type='solid'
)
red_fill = PatternFill(
    start_color='FFFF0000', end_color='FFFF0000', fill_type='solid'
    )

# Iterate through each row and column in the worksheet
for row in ws.iter_rows(
    min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            if cell.value > 0:
                cell.fill = green_fill
            elif cell.value < 0:
                cell.fill = red_fill

# Save the workbook with changes
wb.save("InvestmentCalc.xlsx")
